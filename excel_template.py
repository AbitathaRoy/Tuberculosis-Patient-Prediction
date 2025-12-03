# excel_template.py

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from preprocess import FEATURE_ORDER

def create_template(path, max_rows=5000):

    """
    path: Can be a string filename OR a BytesIO buffer object.
    max_rows: How far down to apply the dropdown menus.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Full ordered column list
    columns = ["PatientName", "PatientID"] + FEATURE_ORDER

    # Write header
    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # ---- Dropdown options for each categorical field ----
    gender_options = ["Female", "Male", "Transgender"]
    diabetes_options = ["Non-diabetic", "Diabetic"]
    micro_options = ["Yes", "No"]
    site_options = ["Pulmonary", "Extra Pulmonary"]
    interstate_options = ["Inter-District", "Inter-State"]
    hiv_options = ["Non-Reactive", "Unknown", "Positive", "Reactive"]
    typeofcase_options = [
        "Retreatment: Recurrent",
        "New",
        "Retreatment: Others",
        "PMDT",
        "Retreatment: Treatment after failure",
        "Retreatment: Treatment after lost to follow up"
    ]

    # Map feature names → option lists
    option_map = {
        "Gender": gender_options,
        "DiabetesStatus": diabetes_options,
        "Microbiologically_Confirmed": micro_options,
        "SiteOfDisease": site_options,
        "Inter-state/Inter-district enrollment": interstate_options,
        "HIV_Status": hiv_options,
        "TypeOfCase": typeofcase_options
    }

    # Create data validation for each categorical column
    # Create data validation for each categorical column
    for col_idx, col_name in enumerate(columns, start=1):
        if col_name in option_map:  # Only attach dropdowns to categorical columns
            options = option_map[col_name]
            # Join options with commas for the list formula
            dv = DataValidation(type="list", formula1=f'"{",".join(options)}"')
            ws.add_data_validation(dv)
            
            # Apply validation to rows 2 → max_rows
            # Using coordinate ensures we get "A2", "B2", etc. correctly
            first_cell = ws.cell(row=2, column=col_idx).coordinate
            last_cell = ws.cell(row=max_rows, column=col_idx).coordinate
            dv.add(f"{first_cell}:{last_cell}")


    # Save Excel file
    # If 'path' is a BytesIO object, openpyxl writes to memory automatically.
    wb.save(path)
